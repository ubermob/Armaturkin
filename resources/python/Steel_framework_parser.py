"""
Simple excel parse util.
Requied boot parameter:
- path to .xlsx file
- boolean value: True for translit rus -> eng

author: Andrey Korneychuk on 05-Jan-22
version: 1.2
"""
import openpyxl
import sys
import collections
from pathlib import Path
from transliterate import translit


def print_head():
    print("Parsing columns")
    string = f"A: {sheet.cell(1, 1).value} - число таких выносок"
    if is_translitable:
        string = translit(string, language_code="ru", reversed=True)
    print(string)

    string = f"E: {sheet.cell(1, 5).value} - число каркасов в одной выноске"
    if is_translitable:
        string = translit(string, language_code="ru", reversed=True)
    print(string)

    string = f"G: {sheet.cell(1, 7).value}"
    if is_translitable:
        string = translit(string, language_code="ru", reversed=True)
    print(string)


def contain_key(value, dictionary):
    return value in dictionary


def cast_to_string(value):
    if type(value) != str:
        value = str(value)
    return value


def cast_to_integer(value):
    if type(value) == str:
        value = int(value)
    elif value == None:
        # Default numeric value for empty cell
        value = 1
    return value


def sum_of_list(list):
    sum = 0
    for element in list:
        sum += element
    return sum


def parse_string_to_bool(s: str):
    return s.lower() in ["true"]


path_to_file = sys.argv[1]
try:
    is_translitable = parse_string_to_bool(sys.argv[2])
except Exception:
    is_translitable = False

excel_file = Path(path_to_file)
sheet = openpyxl.load_workbook(excel_file).active
print_head()
dictionary = {}
row_counter = 1
head_row = True

for row in sheet.iter_rows():
    if head_row:
        head_row = False
        row_counter += 1
        continue
    steel_frame_mark = cast_to_string(row[6].value)  # G cell
    value_is_ready = contain_key(steel_frame_mark, dictionary)
    number = cast_to_integer(row[0].value) * cast_to_integer(row[4].value)  # A and E cells
    if value_is_ready:
        dictionary[steel_frame_mark].append(number)
    else:
        dictionary[steel_frame_mark] = [number]
    print(f"row: {row_counter}, OK, ADD ({steel_frame_mark}: {number})")
    row_counter += 1

dictionary = collections.OrderedDict(sorted(dictionary.items()))

print("\n====S=U=M=M=A=R=Y====\n")

global_sum = 0
for key in dictionary.keys():
    key_sum = sum_of_list(dictionary[key])
    global_sum += key_sum
    print(f"key: {key}, sum: {key_sum}")
print(f"\nGlobal sum: {global_sum}")
